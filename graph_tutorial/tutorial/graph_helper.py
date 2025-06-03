# Copyright (c) Microsoft Corporation.
# Licensed under the MIT License.

import json
import requests
from django.utils.dateparse import parse_datetime
from typing import TYPE_CHECKING, List, Dict, Any
from urllib.parse import quote
import pandas as pd
from io import BytesIO
from .models import TaskNotification
from bs4 import BeautifulSoup
from collections import defaultdict
from openpyxl.utils import get_column_letter
if TYPE_CHECKING:
    from .models import AutoScheduleMeeting
import yaml
GRAPH_URL = 'https://graph.microsoft.com/v1.0'

def get_user(token):
    # Send GET to /me
    user = requests.get(
        f'{GRAPH_URL}/me',
        headers={
          'Authorization': f'Bearer {token}'
        },
        params={
          '$select': 'displayName,mail,mailboxSettings,userPrincipalName'
        })
    # Return the JSON result
    return user.json()

def get_users(token, query=None):
    if not query:
        return []
    query = query.strip()
    if not query:
        raise ValueError("Query parameter is required")

    filter_query = f"startswith(displayName,'{query}') or startswith(mail,'{query}')"
    endpoint = f"{GRAPH_URL}/users?$filter={filter_query}&$select=displayName,mail,userPrincipalName"

    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }

    try:
        response = requests.get(endpoint, headers=headers)
        response.raise_for_status()
        data = response.json()

        if "value" not in data:
            raise ValueError("Invalid response from Graph API")

        contacts = []
        for user in data["value"]:
            email = user.get("mail") or user.get("userPrincipalName")
            if email:
                contacts.append({
                    "name": user.get("displayName", ""),
                    "email": email
                })

        return contacts

    except requests.exceptions.RequestException as e:
        raise RuntimeError(f"Graph API request failed: {e}")



def get_calendar_events(token, start, end, timezone):
    # Set headers
    headers = {
        'Authorization': f'Bearer {token}',
        'Prefer': f'outlook.timezone="{timezone}"'
    }

    # Configure query parameters to
    # modify the results
    query_params = {
        'startDateTime': start,
        'endDateTime': end,
        '$select': 'subject,organizer,start,end',
        '$orderby': 'start/dateTime',
        '$top': '50'
    }

    # Send GET to /me/events
    events = requests.get(f'{GRAPH_URL}/me/calendarview',
        headers=headers,
        params=query_params)

    # Return the JSON result
    return events.json()

def get_meeting_times_slots(token: str, meeting: 'AutoScheduleMeeting', timezone: str = 'UTC') -> List[Dict[str, Any]]:
    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type': 'application/json',
        'Prefer': f'outlook.timezone="{timezone}"'
    }
    attendees_list = json.loads(meeting.attendees)  # 將 JSON 字符串轉換為列表
    attendees_list.append(meeting.host_email)

    body = {
        "attendees": [
            {
                "emailAddress": { "address": email },
                "type": "Required"
            } for email in attendees_list  # 確保將 JSON 字符串轉換為列表
        ],
        "timeConstraint": {
            "timeslots": [
                {
                    "start": {
                        "dateTime": meeting.start_time.replace(tzinfo=None).isoformat(),
                        "timeZone": timezone
                    },
                    "end": {
                        "dateTime": meeting.end_time.replace(tzinfo=None).isoformat(),
                        "timeZone": timezone
                    }
                }
            ]
        },
        "meetingDuration": f"PT{meeting.duration}M"
    }

    response = requests.post(f'{GRAPH_URL}/me/findMeetingTimes', headers=headers, json=body)

    if response.status_code != 200:
        raise Exception(f"Microsoft Graph API Error: {response.status_code} {response.text}")

    data = response.json()

    # Check if there are no available time slots
    if not data.get("meetingTimeSuggestions"):
        raise Exception("No available meeting time slots found in the response.")

    meeting_times = []
    for suggestion in data.get("meetingTimeSuggestions", []):
        slot = suggestion["meetingTimeSlot"]
        start_dt = parse_datetime(slot["start"]["dateTime"]).isoformat()
        end_dt = parse_datetime(slot["end"]["dateTime"]).isoformat()
        meeting_times.append({
            "confidence": suggestion["confidence"],
            "attendeeAvailability": suggestion["attendeeAvailability"],
            "start": start_dt,
            "end": end_dt
        })


    return meeting_times

def get_user_info(token, email):
    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type': 'application/json',
    }
    response = requests.get(f'{GRAPH_URL}/users/{email}', headers = headers)
    user_data = response.json()
    return user_data

def get_all_chats(token):
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }
    chats = []
    url = f"{GRAPH_URL}/me/chats"
    
    while url:
        res = requests.get(url, headers=headers)
        res.raise_for_status()
        data = res.json()
        chats.extend(data.get('value', []))
        url = data.get('@odata.nextLink')

    return chats

# 一次找全部
def get_chat_ids(token, user_ids):
    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type': 'application/json',
    }
    chats = get_all_chats(token)
    if not chats:
        raise Exception(f"Failed to get chats")
    
    chat_ids = []

    for user_id in user_ids:
        matched_chat_id = None

        for chat in chats:
            if chat.get("chatType") != "oneOnOne":
                continue

            # Get chat members
            chat_id = chat.get("id")
            members_resp = requests.get(f"{GRAPH_URL}/chats/{chat_id}/members", headers=headers)
            if members_resp.status_code != 200:
                continue  # skip if can't fetch members

            members = members_resp.json().get("value", [])

            # Find the "other" member's tenantId
            for member in members:
                if member.get("userId") == user_id:
                    matched_chat_id = chat_id
                    break

            if matched_chat_id:
                break  # found matching chat, no need to check more chats

        chat_ids.append(matched_chat_id)

    return chat_ids

def create_card_payload(subject, start_time, end_time, tenant_id, uuid, base_response_url='http:/localhost/webhook/response/'):
    card = {
        "type": "AdaptiveCard",
        "version": "1.4",
        "body": [
            {
                "type": "TextBlock",
                "text": f"📢 會議邀請: {subject}",
                "weight": "Bolder",
                "size": "Medium"
            },
            {
                "type": "TextBlock",
                "text": f"🕒 時間: {start_time} ~ {end_time}"
            }
        ],
        "actions": [
            {
                "type": "Action.OpenUrl",
                "title": "✅ 參加",
                "url": f"{base_response_url}?tenantId={tenant_id}&uuid={str(uuid)}&response=accepted"
            },
            {
                "type": "Action.OpenUrl",
                "title": "❌ 不參加",
                "url": f"{base_response_url}?tenantId={tenant_id}&uuid={str(uuid)}&response=declined"
            }
        ]
    }

    card_payload = {
        "body": {
            "contentType": "html",
            "content": "This message was sent automatically by the Microsoft Automation Tool. <attachment id=\"1\"></attachment>"
        },
        "attachments": [
            {
                "id": "1",
                "contentType": "application/vnd.microsoft.card.adaptive",
                "content": json.dumps(card)
            }
        ]
    }

    return card_payload

with open("oauth_settings.yml","r",encoding="utf-8") as file:
    redirect = yaml.safe_load(file)['redirect']
def inform_attendees(token, meeting: 'AutoScheduleMeeting'):
    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type': 'application/json',
    }

    attendee_responses = meeting.get_attendee_responses()

    for email, data in attendee_responses.items():
        chat_id = data.get('chat_id')
        tenant_id = data.get('tenant_id')

        if chat_id:
            card_payload = create_card_payload(
                subject=meeting.title,
                start_time=meeting.get_candidate_time()['start'],
                end_time=meeting.get_candidate_time()['end'],
                tenant_id=tenant_id,
                uuid = meeting.uuid,
                base_response_url=f"{redirect.replace("/callback","")}/webhook/response/"
            )

            url = f"{GRAPH_URL}/chats/{chat_id}/messages"

            response = requests.post(url, headers=headers, json=card_payload)

            if response.status_code >= 300:
                print(f"[❌] Failed to send card to {email} (chat_id: {chat_id})")
                print(f"Response: {response.status_code} - {response.text}")
            else:
                print(f"[✅] Card sent to {email}")
        else:
            print(f"[⚠️] No chat_id for {email}, skipping")



def create_event(token, subject, start, end, attendees=None, body=None, timezone='UTC'):
    # Create an event object
    # https://docs.microsoft.com/graph/api/resources/event?view=graph-rest-1.0
    new_event = {
        'subject': subject,
        'start': {
            'dateTime': start,
            'timeZone': timezone
        },
        'end': {
            'dateTime': end,
            'timeZone': timezone
        },
        'location': {
            'displayName': "Teams 線上會議",
        },
        'isOnlineMeeting': True,
        'onlineMeetingProvider': "teamsForBusiness",
    }

    if attendees:
        attendee_list = []
        for email in attendees:
            # Create an attendee object
            # https://docs.microsoft.com/graph/api/resources/attendee?view=graph-rest-1.0
            attendee_list.append({
                'type': 'required',
                'emailAddress': { 'address': email }
            })

        new_event['attendees'] = attendee_list

    if body:
        # Create an itemBody object
        # https://docs.microsoft.com/graph/api/resources/itembody?view=graph-rest-1.0
        new_event['body'] = {
            'contentType': 'text',
            'content': body
        }

    # Set headers
    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type': 'application/json'
    }

    response = requests.post(f'{GRAPH_URL}/me/events',
        headers=headers,
        data=json.dumps(new_event))
    if response.status_code != 201:
        print("Failed to create event:", response.status_code, response.text)

    return response

# core
class GraphTeamsClient:
    def __init__(self, access_token):
        self.token = access_token
        self.graph_url = GRAPH_URL
        self.headers = {
            'Authorization': f'Bearer {self.token}',
            'Content-Type': 'application/json',
        }
        self.user_info = self.__get_user_info__()      
        # cached
        self._user_info_cache = {} 
        self._chat_id_cache = {}

    def __get_user_info__(self):
        user = requests.get(f'{self.graph_url}/me',self.headers)
        return user.json()

    def get_user_info(self, email):
        """
        Given a user email, return the user ID. Uses caching to avoid redundant API calls.
        """
        if email in self._user_info_cache:
            # Return cached user info if available
            return self._user_info_cache[email]

        url = f"{GRAPH_URL}/users/{email}"
        response = requests.get(url, headers=self.headers)
        if response.status_code != 200:
            raise Exception(f"Failed to get user ID: {response.status_code} {response.text}")

        user_data = response.json()
        # Cache the user info
        self._user_info_cache[email] = user_data
        return user_data

    def get_chat_id_by_name(self, chat_name):
            """
            Given a chat name, return the chat ID. Uses caching to avoid redundant API calls.
            """
            # Check if the chat ID is already cached
            if chat_name in self._chat_id_cache:
                return self._chat_id_cache[chat_name]

            url = f"{GRAPH_URL}/me/chats"
            while url:
                response = requests.get(url, headers=self.headers)
                if response.status_code != 200:
                    raise Exception(f"Failed to get chats: {response.status_code} {response.text}")

                data = response.json()
                chats = data.get("value", [])
                for chat in chats:
                    if chat.get("topic") == chat_name:
                        chat_id = chat.get("id")
                        # Cache the chat ID
                        self._chat_id_cache[chat_name] = chat_id
                        return chat_id

                # Get the next page of results
                url = data.get("@odata.nextLink")

            raise Exception(f"Chat with name '{chat_name}' not found")

    def send_message_to_chat(self, chat_id, message_payload):
        """
        Send a message to a specific chat.
        """
        url = f"{GRAPH_URL}/chats/{chat_id}/messages"
        response = requests.post(url, headers=self.headers, json=message_payload)
        if response.status_code >= 300:
            raise Exception(f"Failed to send message: {response.status_code} {response.text}")
        return response.json()['id']
    # for scrum usage
    def list_msg_in_chats(self, chat_id):
        """
        List all messages in a chat.
        """
        url = f"{GRAPH_URL}/chats/{chat_id}/messages"
        messages = []

        while url:
            response = requests.get(url, headers=self.headers)
            if response.status_code != 200:
                raise Exception(f"Failed to fetch messages: {response.status_code} {response.text}")

            data = response.json()
            messages.extend(data.get("value", []))
            url = data.get("@odata.nextLink")  # Get the next page of messages, if available

        return messages

# sharepoint automation
# 一份excel 實例一個
class GraphSharePointClient(GraphTeamsClient):
    def __init__(self, access_token, path="Feature to do list+Q&A/[19.10] Mx Feature_to do list+ Q&A.xlsx", site_name="NebulaP8group", drive_name="ScrumSprints", domain="unizyx.sharepoint.com"):
        super().__init__(access_token)
        self.path = quote(path)
        self.domain = domain
        self.site_name = site_name
        self.drive_name = drive_name
        self.site_id = self._get_site_id()
        self.list_id = self._get_list_id()
        self.drive_id = self._get_drive_id()
        self.model = TaskNotification

        # column index for the template sheet
        self.col_tag = {
            "Task": 4, "Owner": 5, "EST_start_BE": 6, "EST_start_FE": 7,
            "EST_days_BE": 8, "EST_days_FE": 9, "spent_days_BE": 10,
            "spent_days_FE": 11, "due_date_BE": 12, "due_date_FE": 13,
            "Note": 14, "MR": 15, "teams_group_name": 16
        }

    def _get(self, url):
        res = requests.get(url, headers=self.headers)
        if res.status_code != 200:
            raise Exception(f"GET failed: {res.status_code} {res.text}")
        return res.json()

    def _patch(self, url, json_payload):
        res = requests.patch(url, headers=self.headers, json=json_payload)
        if res.status_code != 200:
            raise Exception(f"PATCH failed: {res.status_code} {res.text}")
        return res.json()

    def _get_site_id(self):
        url = f"{self.graph_url}/sites/{self.domain}:/sites/{self.site_name}"
        self.site_id = self._get(url).get("id")
        return self.site_id

    def _get_drive_id(self):
        url = f"{self.graph_url}/sites/{self._get_site_id()}/drives"
        for drive in self._get(url)["value"]:
            if drive["name"] == self.drive_name:
                self._drive_id = drive["id"]
                return self._drive_id
        raise Exception(f"Drive {self.drive_name} not found")
    
    def _get_list_id(self, drive_name="ScrumSprints"):
        url = f"{self.graph_url}/sites/{self._get_site_id()}/lists"
        for lst in self._get(url)["value"]:
            if lst["displayName"] == drive_name:
                self._list_id = lst.get("id")
                return self._list_id
        raise Exception(f"List with name '{drive_name}' not found")

    # for download file usage
    def _build_drive_url(self):
        return f"{self.graph_url}/sites/{self.site_id}/drives/{self.drive_id}/root:/{self.path}"

    # for write file usage
    def _build_list_url(self):
        return f"{self.graph_url}/sites/{self.site_id}/lists/{self.list_id}/drive/root:/{self.path}"
    
    def _build_excel_range_url(self, sheet, address):
        return f"{self._build_list_url()}:/workbook/worksheets('{sheet}')/range(address='{address}')"
    
    # return a dict with sheet name as key and DataFrame as value
    def _download_excel_as_df(self, sheet_name=None, file_type="xlsx"):
        url = f"{self._build_drive_url()}:/content"
        res = requests.get(url, headers=self.headers)
        if res.status_code != 200:
            raise Exception(f"Download failed: {res.status_code} {res.text}")
        if file_type == "csv":
            return pd.read_csv(BytesIO(res.content), sheet_name)
        elif file_type in ["xls", "xlsx"]:
            return pd.read_excel(BytesIO(res.content), sheet_name)
        else:
            raise ValueError("Unsupported file type")

    def _write_cell(self, uuid, values):
        task = self.model.objects.get(uuid=uuid)
        url = self._build_excel_range_url(task.sheet_name, task.field_address)
        payload = {"values": values}
        self._patch(url, payload)
        print("✅ Updated")
        task.replied = True
        task.save()
    
    def _create_notify_item(self, context: dict, reason: str, field: str):
        try:
            chat_id = self.get_chat_id_by_name(context["teams_group_name"])
        except Exception as e:
            print(f"no chat id\n{context}")
            return
        user_info = self.get_user_info(context['owner'])
        # 發送 Teams 通知
        payload = self._create_mention_message_payload(
            context,
            reason
        )
        msg_id = self.send_message_to_chat(chat_id, payload)

        # 查詢條件
        lookup = {
            "sheet_name": context["sheet_name"],
            "row": context["row_idx"],
            "reason": reason
        }

        # 預設欄位（建新時使用）
        defaults = {
            "task": context["task"],
            "owner_id": user_info["id"],
            "owner_email": context["owner"],
            "owner_name": user_info["displayName"],
            "teams_group_id": chat_id,
            "teams_group_name": context["teams_group_name"],
            "field_address": field,
            "msg_id": [msg_id], 
        }

        obj, created = self.model.objects.get_or_create(defaults=defaults, **lookup)

        if not created:
            # ✅ 更新現有資料並 append msg_id（避免重複）
            if msg_id not in obj.msg_id:
                obj.msg_id.append(msg_id)

            # 同步更新其餘欄位（視情況保留）
            for key, value in defaults.items():
                if key != "msg_id":
                    setattr(obj, key, value)

            obj.save()


    def _process_sheet(self, df, sheet_name):
        for row_idx, row in df.iterrows():
            try:
                task = row.iloc[self.col_tag["Task"]]
                owner = row.iloc[self.col_tag["Owner"]]
                teams_group_name = row.iloc[self.col_tag["teams_group_name"]]
            except Exception as e:
                # print(f"{sheet_name} is not template format")
                continue
            # 被放入notify的條件
            if pd.isna(task) or pd.isna(owner) or not isinstance(owner, str) or "@" not in owner or pd.isna(teams_group_name):
                continue
            context = {
                "sheet_name": sheet_name,
                "row_idx": row_idx,
                "task": task,
                "owner": owner,
                "teams_group_name": teams_group_name,
            }

            # add for more trigger conditions
            #　一個條件一則通知
            # 檢查 start date
            est_start_be = row.iloc[self.col_tag["EST_start_BE"]]
            if pd.isna(est_start_be):
                col = get_column_letter(self.col_tag["EST_start_BE"] + 1)
                self._create_notify_item(
                    context,
                    reason="Estimate start date BE is missing",
                    field=f"{col}{row_idx + 2}"
                )

            est_start_fe = row.iloc[self.col_tag["EST_start_FE"]]
            if pd.isna(est_start_fe):
                col = get_column_letter(self.col_tag["EST_start_FE"] + 1)
                self._create_notify_item(
                    context,
                    reason="Estimate start date FE is missing",
                    field=f"{col}{row_idx + 2}"
                )
            # 檢查 due date 為空
            due_date_be = row.iloc[self.col_tag["due_date_BE"]]
            if pd.isna(due_date_be):
                col = get_column_letter(self.col_tag["due_date_BE"] + 1)
                self._create_notify_item(
                    context,
                    reason="Due date BE is missing",
                    field=f"{col}{row_idx + 2}"
                )
            due_date_fe = row.iloc[self.col_tag["due_date_FE"]]
            if pd.isna(due_date_fe):
                col = get_column_letter(self.col_tag["due_date_FE"] + 1)
                self._create_notify_item(
                    context,
                    reason="Due date FE is missing",
                    field=f"{col}{row_idx + 2}"
                )

            # 檢查 estimated start day 與今天日期差一天
            today = pd.Timestamp.now().normalize()
            if not pd.isna(est_start_be) and abs((est_start_be - today).days) == 1:
                col = get_column_letter(self.col_tag["EST_start_BE"] + 1)
                self._create_notify_item(
                    context,
                    reason="Estimated start date BE is within one day of today",
                    field=f"{col}{row_idx + 2}"
                )
            if not pd.isna(est_start_fe) and abs((est_start_fe - today).days) == 1:
                col = get_column_letter(self.col_tag["EST_start_FE"] + 1)
                self._create_notify_item(
                    context,
                    reason="Estimated start date FE is within one day of today",
                    field=f"{col}{row_idx + 2}"
                )

            # 檢查 due start day 與今天日期差一天
            if not pd.isna(due_date_be) and abs((due_date_be - today).days) == 1:
                col = get_column_letter(self.col_tag["due_date_BE"] + 1)
                self._create_notify_item(
                    context,
                    reason="Due date BE is within one day of today",
                    field=f"{col}{row_idx + 2}"
                )
            if not pd.isna(due_date_fe) and abs((due_date_fe - today).days) == 1:
                col = get_column_letter(self.col_tag["due_date_FE"] + 1)
                self._create_notify_item(
                    context,
                    reason="Due date FE is within one day of today",
                    field=f"{col}{row_idx + 2}"
                )

    def _create_mention_message_payload(self, context, reason):
        """
        Create a message payload with a mention for a specific user.

        Args:
            context (dict): Includes 'owner', 'sheet_name', 'task', etc.
            reason (str): The reason for the message.

        Returns:
            dict: The payload for sending the message.
        """
        user_info = self.get_user_info(context['owner'])

        # Construct the mention object
        mention = {
            "id": 0,  # Must match <at id="0"> in content
            "mentionText": user_info['displayName'],
            "mentioned": {
                "user": {
                    "id": user_info['id'],
                    "displayName": user_info['displayName']
                }
            }
        }

        # Construct the message payload
        payload = {
            "body": {
                "contentType": "html",
                "content": (
                    f"<div>"
                    f"<p>👋 <at id=\"0\">{user_info['displayName']}</at>, please reply to this message.</p>"
                    f"<p>💬 <i>(Your reply will be automatically recorded to SharePoint)</i></p>"
                    f"<p>📄 <b>Sheet:</b> {context.get('sheet_name', 'N/A')}</p>"
                    f"<p>📝 <b>Task:</b> {context.get('task', 'N/A')}</p>"
                    f"<p>⚠️ <b>Reason:</b> {reason}</p>"
                    f"</div>"
                )
            },
            "mentions": [mention]
        }

        return payload



    def _search_message_reference(self, messages, user_id, msg_id):
        """
        Search from cached messages list (not API call) for a reply that:
        - is from the given user
        - is a messageReference type
        - references the given msg_id
        """
        for message in messages:
            attachments = message.get("attachments", [])
            if not attachments:
                continue

            if (
                message.get("from", {}).get("user", {}).get("id") == user_id and
                attachments[0].get("contentType") == "messageReference" and
                attachments[0].get("id") == msg_id
            ):
                soup = BeautifulSoup(message['body']['content'], "html.parser")

                for emoji_tag in soup.find_all("emoji"):
                    if emoji_tag.has_attr("alt"):
                        emoji_tag.replace_with(emoji_tag["alt"])


                text = soup.get_text(separator=' ', strip=True)
                return text

        return None
    # routine
    def scan_routine(self, sheet_name="automation_test"):
        """
        Process the specified sheet or all sheets in the Excel file and store the results in the database.
        inform task owner on temas
        Args:
            sheet_name (str or None): The name of the sheet to process. If None, all sheets will be processed.
        Returns:
            None
        """
        # 每次都重新抓取最新資料來生成notify item
        self.model.objects.all().delete()
        sheets = self._download_excel_as_df(sheet_name=sheet_name)

        if sheet_name is not None:
            # 處理單一工作表
            self._process_sheet(sheets, sheet_name)
        else:
            # 處理多個工作表
            for name, df in sheets.items():
                self._process_sheet(df, name)
    # polling
    def polling_task_pool(self):
        
        # 1. Load all notification records
        notifications = self.model.objects.all()

        # 2. Group by chat_id
        chat_groups = defaultdict(list)
        # 將每個 item 的完整資訊加入對應的 chat_groups
        for item in notifications:
            chat_groups[item.teams_group_id].append({
                "uuid": item.uuid,
                "owner_id": item.owner_id,
                "msg_id": item.msg_id,
                "task": item.task
            })

        # 3. Iterate each chat group and fetch messages once
        for chat_id, items in chat_groups.items():
            try:
                messages = self.list_msg_in_chats(chat_id)
            except Exception as e:
                print(f"⚠️ Failed to fetch messages for chat {chat_id}: {e}")
                continue

            # 4. Search for replies matching user_id and msg_id in current chat
            for item in items:
                try:
                    user_id = item['owner_id']
                    for mid in item['msg_id']:
                        content = self._search_message_reference(messages, user_id, mid)
                        if content:
                            self._write_cell(item['uuid'], content)
                            print(f"📝 Replied content written for task {item['task']}")
                            break  # only process first found reply
                except Exception as e:
                    print(f"❌ Error processing task {item['task']}: {e}")

#/* spell-checker: disable */
# Basic lookup for mapping Windows time zone identifiers to
# IANA identifiers
# Mappings taken from
# https://github.com/unicode-org/cldr/blob/master/common/supplemental/windowsZones.xml
zone_mappings = {
    'Dateline Standard Time': 'Etc/GMT+12',
    'UTC-11': 'Etc/GMT+11',
    'Aleutian Standard Time': 'America/Adak',
    'Hawaiian Standard Time': 'Pacific/Honolulu',
    'Marquesas Standard Time': 'Pacific/Marquesas',
    'Alaskan Standard Time': 'America/Anchorage',
    'UTC-09': 'Etc/GMT+9',
    'Pacific Standard Time (Mexico)': 'America/Tijuana',
    'UTC-08': 'Etc/GMT+8',
    'Pacific Standard Time': 'America/Los_Angeles',
    'US Mountain Standard Time': 'America/Phoenix',
    'Mountain Standard Time (Mexico)': 'America/Chihuahua',
    'Mountain Standard Time': 'America/Denver',
    'Central America Standard Time': 'America/Guatemala',
    'Central Standard Time': 'America/Chicago',
    'Easter Island Standard Time': 'Pacific/Easter',
    'Central Standard Time (Mexico)': 'America/Mexico_City',
    'Canada Central Standard Time': 'America/Regina',
    'SA Pacific Standard Time': 'America/Bogota',
    'Eastern Standard Time (Mexico)': 'America/Cancun',
    'Eastern Standard Time': 'America/New_York',
    'Haiti Standard Time': 'America/Port-au-Prince',
    'Cuba Standard Time': 'America/Havana',
    'US Eastern Standard Time': 'America/Indianapolis',
    'Turks And Caicos Standard Time': 'America/Grand_Turk',
    'Paraguay Standard Time': 'America/Asuncion',
    'Atlantic Standard Time': 'America/Halifax',
    'Venezuela Standard Time': 'America/Caracas',
    'Central Brazilian Standard Time': 'America/Cuiaba',
    'SA Western Standard Time': 'America/La_Paz',
    'Pacific SA Standard Time': 'America/Santiago',
    'Newfoundland Standard Time': 'America/St_Johns',
    'Tocantins Standard Time': 'America/Araguaina',
    'E. South America Standard Time': 'America/Sao_Paulo',
    'SA Eastern Standard Time': 'America/Cayenne',
    'Argentina Standard Time': 'America/Buenos_Aires',
    'Greenland Standard Time': 'America/Godthab',
    'Montevideo Standard Time': 'America/Montevideo',
    'Magallanes Standard Time': 'America/Punta_Arenas',
    'Saint Pierre Standard Time': 'America/Miquelon',
    'Bahia Standard Time': 'America/Bahia',
    'UTC-02': 'Etc/GMT+2',
    'Azores Standard Time': 'Atlantic/Azores',
    'Cape Verde Standard Time': 'Atlantic/Cape_Verde',
    'UTC': 'Etc/GMT',
    'GMT Standard Time': 'Europe/London',
    'Greenwich Standard Time': 'Atlantic/Reykjavik',
    'Sao Tome Standard Time': 'Africa/Sao_Tome',
    'Morocco Standard Time': 'Africa/Casablanca',
    'W. Europe Standard Time': 'Europe/Berlin',
    'Central Europe Standard Time': 'Europe/Budapest',
    'Romance Standard Time': 'Europe/Paris',
    'Central European Standard Time': 'Europe/Warsaw',
    'W. Central Africa Standard Time': 'Africa/Lagos',
    'Jordan Standard Time': 'Asia/Amman',
    'GTB Standard Time': 'Europe/Bucharest',
    'Middle East Standard Time': 'Asia/Beirut',
    'Egypt Standard Time': 'Africa/Cairo',
    'E. Europe Standard Time': 'Europe/Chisinau',
    'Syria Standard Time': 'Asia/Damascus',
    'West Bank Standard Time': 'Asia/Hebron',
    'South Africa Standard Time': 'Africa/Johannesburg',
    'FLE Standard Time': 'Europe/Kiev',
    'Israel Standard Time': 'Asia/Jerusalem',
    'Kaliningrad Standard Time': 'Europe/Kaliningrad',
    'Sudan Standard Time': 'Africa/Khartoum',
    'Libya Standard Time': 'Africa/Tripoli',
    'Namibia Standard Time': 'Africa/Windhoek',
    'Arabic Standard Time': 'Asia/Baghdad',
    'Turkey Standard Time': 'Europe/Istanbul',
    'Arab Standard Time': 'Asia/Riyadh',
    'Belarus Standard Time': 'Europe/Minsk',
    'Russian Standard Time': 'Europe/Moscow',
    'E. Africa Standard Time': 'Africa/Nairobi',
    'Iran Standard Time': 'Asia/Tehran',
    'Arabian Standard Time': 'Asia/Dubai',
    'Astrakhan Standard Time': 'Europe/Astrakhan',
    'Azerbaijan Standard Time': 'Asia/Baku',
    'Russia Time Zone 3': 'Europe/Samara',
    'Mauritius Standard Time': 'Indian/Mauritius',
    'Saratov Standard Time': 'Europe/Saratov',
    'Georgian Standard Time': 'Asia/Tbilisi',
    'Volgograd Standard Time': 'Europe/Volgograd',
    'Caucasus Standard Time': 'Asia/Yerevan',
    'Afghanistan Standard Time': 'Asia/Kabul',
    'West Asia Standard Time': 'Asia/Tashkent',
    'Ekaterinburg Standard Time': 'Asia/Yekaterinburg',
    'Pakistan Standard Time': 'Asia/Karachi',
    'Qyzylorda Standard Time': 'Asia/Qyzylorda',
    'India Standard Time': 'Asia/Calcutta',
    'Sri Lanka Standard Time': 'Asia/Colombo',
    'Nepal Standard Time': 'Asia/Katmandu',
    'Central Asia Standard Time': 'Asia/Almaty',
    'Bangladesh Standard Time': 'Asia/Dhaka',
    'Omsk Standard Time': 'Asia/Omsk',
    'Myanmar Standard Time': 'Asia/Rangoon',
    'SE Asia Standard Time': 'Asia/Bangkok',
    'Altai Standard Time': 'Asia/Barnaul',
    'W. Mongolia Standard Time': 'Asia/Hovd',
    'North Asia Standard Time': 'Asia/Krasnoyarsk',
    'N. Central Asia Standard Time': 'Asia/Novosibirsk',
    'Tomsk Standard Time': 'Asia/Tomsk',
    'China Standard Time': 'Asia/Shanghai',
    'North Asia East Standard Time': 'Asia/Irkutsk',
    'Singapore Standard Time': 'Asia/Singapore',
    'W. Australia Standard Time': 'Australia/Perth',
    'Taipei Standard Time': 'Asia/Taipei',
    'Ulaanbaatar Standard Time': 'Asia/Ulaanbaatar',
    'Aus Central W. Standard Time': 'Australia/Eucla',
    'Transbaikal Standard Time': 'Asia/Chita',
    'Tokyo Standard Time': 'Asia/Tokyo',
    'North Korea Standard Time': 'Asia/Pyongyang',
    'Korea Standard Time': 'Asia/Seoul',
    'Yakutsk Standard Time': 'Asia/Yakutsk',
    'Cen. Australia Standard Time': 'Australia/Adelaide',
    'AUS Central Standard Time': 'Australia/Darwin',
    'E. Australia Standard Time': 'Australia/Brisbane',
    'AUS Eastern Standard Time': 'Australia/Sydney',
    'West Pacific Standard Time': 'Pacific/Port_Moresby',
    'Tasmania Standard Time': 'Australia/Hobart',
    'Vladivostok Standard Time': 'Asia/Vladivostok',
    'Lord Howe Standard Time': 'Australia/Lord_Howe',
    'Bougainville Standard Time': 'Pacific/Bougainville',
    'Russia Time Zone 10': 'Asia/Srednekolymsk',
    'Magadan Standard Time': 'Asia/Magadan',
    'Norfolk Standard Time': 'Pacific/Norfolk',
    'Sakhalin Standard Time': 'Asia/Sakhalin',
    'Central Pacific Standard Time': 'Pacific/Guadalcanal',
    'Russia Time Zone 11': 'Asia/Kamchatka',
    'New Zealand Standard Time': 'Pacific/Auckland',
    'UTC+12': 'Etc/GMT-12',
    'Fiji Standard Time': 'Pacific/Fiji',
    'Chatham Islands Standard Time': 'Pacific/Chatham',
    'UTC+13': 'Etc/GMT-13',
    'Tonga Standard Time': 'Pacific/Tongatapu',
    'Samoa Standard Time': 'Pacific/Apia',
    'Line Islands Standard Time': 'Pacific/Kiritimati'
}

def get_iana_from_windows(windows_tz_name):
    if windows_tz_name in zone_mappings:
        return zone_mappings[windows_tz_name]

    # Assume if not found value is
    # already an IANA name
    return windows_tz_name
